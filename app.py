import streamlit as st
import pandas as pd
import datetime
import io
import csv
from openpyxl.utils import get_column_letter

# Configuración de la página para usar todo el ancho disponible
st.set_page_config(layout="wide", page_title="Dice debe Decir - Aplicación de Gasto")

# ----- FUNCIONES DE CARGA DE DATOS -----

@st.cache_data
def load_base_data():
    """Carga la base de datos desde el archivo Excel."""
    return pd.read_excel("BASE DE DATOS.xlsx")

@st.cache_data
def load_conversion_factors():
    """Carga los factores de conversión desde el archivo CSV."""
    conversion = {}
    with open("factores_conversion.csv", newline='', encoding="latin-1") as csvfile:
        reader = csv.reader(csvfile, delimiter="\t")
        headers = next(reader)
        year_headers = [int(h.strip()) for h in headers[1:]]
        for row in reader:
            base_year = int(row[0].strip())
            conversion[base_year] = {}
            for idx, cell in enumerate(row[1:]):
                try:
                    value = float(cell.strip().replace(",", "."))
                except ValueError:
                    value = None
                conversion[base_year][year_headers[idx]] = value
    return conversion

# ----- FORMATO NUMÉRICO PERSONALIZADO -----

def format_number_custom(x):
    """
    Convierte un número a entero sin decimales y lo formatea con puntos como separador de miles.
    Ejemplo: 9182936 se mostrará como "9.182.936" y 0 como "0".
    """
    try:
        return f"{int(round(x)):,}".replace(",", ".")
    except Exception:
        return x

def style_df_contabilidad(df):
    """
    Devuelve un objeto Styler que aplica el formato contable:
      - Los valores numéricos se formatean sin decimales y con puntos como separador de miles.
      - Toda la tabla se alinea a la izquierda.
    """
    styler = df.style.format(lambda x: format_number_custom(x) if isinstance(x, (int, float)) else x)
    styler = styler.set_properties(**{'text-align': 'left'})
    # Alinear los encabezados a la izquierda
    styler = styler.set_table_styles([{'selector': 'th', 'props': [('text-align', 'left')]}])
    return styler

# ----- FUNCION PARA AGREGAR TOTALES (fila y columna) -----

def append_totals(df):
    """
    Agrega una columna "Total" (suma de las columnas numéricas) a cada fila y
    añade una fila final "Total" con la suma de cada columna numérica.
    Las columnas no numéricas quedan vacías en la fila total.
    """
    df = df.copy()
    numeric_cols = df.select_dtypes(include=["number"]).columns
    df["Total"] = df[numeric_cols].sum(axis=1)
    total_row = df[numeric_cols].sum(axis=0)
    total_row["Total"] = total_row.sum()
    for col in df.columns.difference(numeric_cols):
        total_row[col] = ""
    total_row.name = "Total"
    df = pd.concat([df, total_row.to_frame().T])
    return df

# ----- FILTRADO Y GENERACIÓN DE LA PLANILLA -----

def get_filtered_data(df_base, codigo_bip, etapa, anio_termino):
    codigo_bip_norm = str(codigo_bip).strip().upper()
    etapa_norm = str(etapa).strip().upper()
    df_filtered = df_base[
        (df_base["CODIGO BIP"].astype(str).str.strip().str.upper() == codigo_bip_norm) &
        (df_base["ETAPA"].astype(str).str.strip().str.upper() == etapa_norm)
    ]
    if df_filtered.empty:
        st.error("No se encontraron datos para el CODIGO BIP y ETAPA seleccionados.")
        return None, None, None
    df_filtered.columns = [str(col).strip() for col in df_filtered.columns]
    expense_cols = [col for col in df_filtered.columns if col.isdigit() and 2011 <= int(col) <= 2024]
    df_grouped = df_filtered.groupby("ITEMS")[expense_cols].sum()
    sorted_years = sorted([int(col) for col in expense_cols])
    start_year = None
    for y in sorted_years:
        col = str(y)
        col_sum = df_grouped[col].sum() if col in df_grouped.columns else 0
        if col_sum > 0:
            start_year = y
            break
    if start_year is None:
        st.error("No se encontró gasto inicial en los datos.")
        return None, None, None
    if anio_termino < start_year:
        st.error("El AÑO DE TERMINO debe ser mayor o igual al año de inicio del gasto.")
        return None, None, None
    global_years = list(range(start_year, anio_termino + 1))
    cols = [str(y) for y in global_years]
    for col in cols:
        if col not in df_grouped.columns:
            df_grouped[col] = 0
    df_grouped = df_grouped[cols].sort_index()
    return df_grouped, global_years, df_filtered

def compute_conversion_table(original_df, global_years, conversion_factors, target_conversion_year):
    conv_df = original_df.copy().astype(float)
    for col in conv_df.columns:
        year = int(col)
        base_key = year if year in conversion_factors else max(conversion_factors.keys())
        available_years = sorted(conversion_factors[base_key].keys())
        target_year_use = target_conversion_year if target_conversion_year <= available_years[-1] else available_years[-1]
        factor = conversion_factors[base_key][target_year_use]
        conv_df[col] = (conv_df[col] * factor) / 1000.0
    return conv_df

def compute_programming_table(original_df, global_years, conversion_factors, target_prog_year):
    start_year = global_years[0]
    if target_prog_year >= start_year:
        st.error("El año de conversión para la Programación debe ser menor que el año de inicio.")
        return None
    prog_df = original_df.copy().astype(float)
    for col in prog_df.columns:
        year = int(col)
        base_key = year if year in conversion_factors else max(conversion_factors.keys())
        available_years = sorted(conversion_factors[base_key].keys())
        target_use = available_years[0] if target_prog_year < available_years[0] else target_prog_year
        factor = conversion_factors[base_key][target_use]
        prog_df[col] = (prog_df[col] * factor) / 1000.0
    return prog_df

def compute_cuadro_extra(conv_df, global_years):
    if global_years is None or len(global_years) == 0:
        st.error("No se definieron años globales para calcular el cuadro extra.")
        return pd.DataFrame()
    current_year = datetime.datetime.now().year
    extra_data = []
    for item in conv_df.index:
        pagado = sum(conv_df.loc[item, str(y)] for y in global_years if y <= (current_year - 1))
        sol2025 = conv_df.loc[item, str(current_year)] if str(current_year) in conv_df.columns else 0
        sol_siguientes = sum(conv_df.loc[item, str(y)] for y in global_years if y > current_year)
        total = pagado + sol2025 + sol_siguientes
        extra_data.append({
            "Fuente": "F.N.D.R.",
            "Asignación Presupuestaria": item,
            "Moneda": "M$",
            "Pagado al 31/12/2024": pagado,
            "Solicitado para el año 2025": sol2025,
            "Solicitado años siguientes": sol_siguientes,
            "Costo Total": total
        })
    extra_df = pd.DataFrame(extra_data).set_index("Asignación Presupuestaria")
    return extra_df

def export_to_excel(original_df, conv_df, extra_df, prog_df, selected_codigo_bip):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        original_df.to_excel(writer, sheet_name="Gasto Real", startrow=2)
        conv_df.to_excel(writer, sheet_name="Conversión", startrow=2)
        extra_df.to_excel(writer, sheet_name="Cuadro Extra", startrow=2)
        if prog_df is not None:
            prog_df.to_excel(writer, sheet_name="Programación", startrow=2)
        workbook = writer.book
        from openpyxl.styles import Font, Alignment
        title_font = Font(bold=True, size=14)
        center_alignment = Alignment(horizontal="center")
        sheets = {
            "Gasto Real": original_df,
            "Conversión": conv_df,
            "Cuadro Extra": extra_df,
            "Programación": prog_df if prog_df is not None else pd.DataFrame()
        }
        for sheet_name, df in sheets.items():
            if sheet_name not in writer.sheets:
                continue
            worksheet = writer.sheets[sheet_name]
            ncols = df.shape[1] + 1 if sheet_name in ["Gasto Real", "Conversión", "Programación"] else df.shape[1]
            last_col_letter = get_column_letter(ncols)
            worksheet.merge_cells(f"A1:{last_col_letter}1")
            title_text = f"Proyecto: {selected_codigo_bip}"
            cell = worksheet["A1"]
            cell.value = title_text
            cell.font = title_font
            cell.alignment = center_alignment
            for col in worksheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[col_letter].width = adjusted_width
        writer.save()
    return output.getvalue()

def main():
    st.title("Dice debe Decir - Aplicación de Gasto")
    df_base = load_base_data()
    conversion_factors = load_conversion_factors()
    
    st.sidebar.header("Filtrar Datos")
    codigo_bip_list = sorted(df_base["CODIGO BIP"].dropna().unique().tolist())
    selected_codigo_bip = st.sidebar.selectbox("Seleccione el CODIGO BIP:", codigo_bip_list)
    etapa_list = sorted(df_base["ETAPA"].dropna().unique().tolist())
    selected_etapa = st.sidebar.selectbox("Seleccione la ETAPA:", etapa_list)
    anio_termino = st.sidebar.number_input("Ingrese el AÑO DE TERMINO del proyecto:",
                                           min_value=2011, max_value=2100, value=2024, step=1)
    
    if st.sidebar.button("Generar Planilla"):
        df_grouped, global_years, df_filtered = get_filtered_data(df_base, selected_codigo_bip, selected_etapa, anio_termino)
        if df_grouped is None:
            return
        
        nombre_proyecto = df_filtered["NOMBRE"].iloc[0] if "NOMBRE" in df_filtered.columns else "Proyecto sin nombre"
        with st.container():
            st.markdown(
                f"""
                <div style="padding:5px; background-color:#f0f0f0; border:1px solid #ccc; border-radius:5px; font-size:14px; margin-bottom:10px;">
                    <b>Proyecto:</b> {nombre_proyecto} &nbsp;&nbsp;&nbsp;
                    <b>Etapa:</b> {selected_etapa} &nbsp;&nbsp;&nbsp;
                    <b>Código BIP:</b> {selected_codigo_bip}
                </div>
                """, unsafe_allow_html=True
            )
        
        # Sección 1: Gasto Real no Ajustado (editor interactivo)
        st.markdown("### Gasto Real no Ajustado")
        st.write("Edite los valores según corresponda:")
        if hasattr(st, "data_editor"):
            edited_original_df = st.data_editor(df_grouped, key="original_editor")
        else:
            edited_original_df = st.experimental_data_editor(df_grouped, key="original_editor")
        
        # Sección 1.2: Tabla con Totales (segunda visualización)
        st.markdown("### Gasto Real no Ajustado Cuadro Completo")
        original_df_totals = append_totals(edited_original_df)
        st.table(style_df_contabilidad(original_df_totals))
        
        # Sección 2: Conversión a Moneda Pesos (M$)
        st.markdown("### Conversión a Moneda Pesos (M$)")
        target_conversion_year = st.number_input("Convertir a año:",
                                                   min_value=2011, max_value=2100,
                                                   value=datetime.datetime.now().year, step=1, key="conv_year")
        conv_df = compute_conversion_table(edited_original_df, global_years, conversion_factors, target_conversion_year)
        conv_df_totals = append_totals(conv_df)
        st.table(style_df_contabilidad(conv_df_totals))
        
        # Sección 3: Cuadro Extra (sin columna "Total")
        st.markdown("### Cuadro Extra")
        extra_df = compute_cuadro_extra(conv_df, global_years)
        st.table(style_df_contabilidad(extra_df))
        
        # Sección 4: Programación en Moneda Original
        st.markdown("### Programación en Moneda Original")
        target_prog_year = st.number_input("Convertir a año (Programación):",
                                             min_value=1900, max_value=2100,
                                             value=2010, step=1, key="prog_year")
        prog_df = compute_programming_table(edited_original_df, global_years, conversion_factors, target_prog_year)
        if prog_df is not None:
            prog_df_totals = append_totals(prog_df)
            st.table(style_df_contabilidad(prog_df_totals))
        
        st.markdown("### Exportar a Excel")
        if st.button("Exportar a Excel"):
            excel_data = export_to_excel(edited_original_df, conv_df, extra_df, prog_df, selected_codigo_bip)
            st.download_button(label="Descargar Excel", data=excel_data,
                               file_name="exported_data.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == '__main__':
    main()
